"""任意文件 → PDF 的归一化。

打印路径只有一条，所有格式先变成 PDF，所以这里错了后面全错。
Office COM 那条路要装 Office 才能跑，标了 needs_office。
"""

from __future__ import annotations

import pymupdf
import pytest
from PIL import Image

from shop_print.core import convert
from shop_print.core.enhance import MODE_TEXT, EnhanceOptions
from shop_print.core.errors import ShopPrintError
from shop_print.texts import ErrorKind

from .synth import photographed_document


def 存图(path, array, dpi: tuple[int, int] | None = None):
    image = Image.fromarray(array[:, :, 0] if array.ndim == 3 else array, mode="L")
    image.save(path, dpi=dpi) if dpi else image.save(path)
    return path


@pytest.fixture
def 拍照样张(tmp_path):
    return 存图(tmp_path / "拍照的合同.png", photographed_document())


# ── 分类 ────────────────────────────────────────────────────────
@pytest.mark.parametrize(
    ("name", "kind"),
    [
        ("a.pdf", "pdf"),
        ("a.PDF", "pdf"),
        ("照片.jpg", "image"),
        ("照片.JPEG", "image"),
        ("合同.docx", "office"),
        ("表格.xls", "office"),
        ("说明.txt", "text"),
        ("视频.mp4", "unsupported"),
        ("没有扩展名", "unsupported"),
    ],
)
def test_按扩展名分类(name: str, kind: str) -> None:
    assert convert.classify(name) == kind
    assert convert.is_supported(name) is (kind != "unsupported")


def test_不支持的类型报的是人话() -> None:
    with pytest.raises(ShopPrintError) as caught:
        convert.to_pdf("顾客发来的.mp4")
    assert caught.value.kind is ErrorKind.FILE_UNSUPPORTED
    assert "请让顾客发成 PDF 或者图片" in caught.value.friendly


def test_文件不存在报文件坏了(tmp_path) -> None:
    with pytest.raises(ShopPrintError) as caught:
        convert.to_pdf(tmp_path / "根本没有这个.pdf")
    assert caught.value.kind is ErrorKind.FILE_BROKEN


# ── 图片 → PDF ──────────────────────────────────────────────────
def test_图片转出A4一页(拍照样张, tmp_path) -> None:
    out = convert.images_to_pdf([拍照样张], tmp_path / "out.pdf", convert.ConvertOptions())
    with pymupdf.open(out) as doc:
        assert doc.page_count == 1
        rect = doc[0].rect
        assert abs(rect.width - 595.276) < 1  # A4 纵向
        assert abs(rect.height - 841.89) < 1


def test_多张图合成多页(拍照样张, tmp_path) -> None:
    second = 存图(tmp_path / "第二张.png", photographed_document(seed=7))
    out = convert.images_to_pdf([拍照样张, second], tmp_path / "两张.pdf", convert.ConvertOptions())
    with pymupdf.open(out) as doc:
        assert doc.page_count == 2


def test_横图自动用横向纸省纸(tmp_path) -> None:
    # 截出来的是 600 高 × 1200 宽，本来就是横的
    横图 = 存图(tmp_path / "横的.png", photographed_document()[:600, :])
    out = convert.images_to_pdf(
        [横图], tmp_path / "横向.pdf", convert.ConvertOptions(auto_orient=True)
    )
    with pymupdf.open(out) as doc:
        assert doc[0].rect.width > doc[0].rect.height


def test_关掉自动方向就一直用纵向(tmp_path) -> None:
    横图 = 存图(tmp_path / "横的.png", photographed_document()[:600, :])
    out = convert.images_to_pdf(
        [横图], tmp_path / "纵向.pdf", convert.ConvertOptions(auto_orient=False)
    )
    with pymupdf.open(out) as doc:
        assert doc[0].rect.width < doc[0].rect.height


def test_适应纸张会留边(拍照样张, tmp_path) -> None:
    out = convert.images_to_pdf(
        [拍照样张], tmp_path / "fit.pdf", convert.ConvertOptions(fit=convert.FIT_FIT)
    )
    with pymupdf.open(out) as doc:
        page = doc[0]
        box = page.get_image_info()[0]["bbox"]
        assert box[0] > 1 and box[1] > 1  # 左上留了边
        assert box[2] < page.rect.width - 1


def test_铺满会盖住整页(拍照样张, tmp_path) -> None:
    out = convert.images_to_pdf(
        [拍照样张], tmp_path / "fill.pdf", convert.ConvertOptions(fit=convert.FIT_FILL)
    )
    with pymupdf.open(out) as doc:
        page = doc[0]
        box = page.get_image_info()[0]["bbox"]
        assert box[0] <= 0.5 or box[1] <= 0.5  # 至少一个方向铺到边（另一方向溢出被裁）
        assert box[2] >= page.rect.width - 0.5 or box[3] >= page.rect.height - 0.5


def test_原尺寸放不下时退回适应纸张而不是裁掉内容(tmp_path) -> None:
    """按 72dpi 标注的大图，原尺寸远超 A4。宁可缩小，绝不能裁掉字。"""
    大图 = 存图(tmp_path / "大.png", photographed_document(), dpi=(72, 72))
    out = convert.images_to_pdf(
        [大图], tmp_path / "actual.pdf", convert.ConvertOptions(fit=convert.FIT_ACTUAL)
    )
    with pymupdf.open(out) as doc:
        page = doc[0]
        box = page.get_image_info()[0]["bbox"]
        assert box[0] >= 0 and box[3] <= page.rect.height + 0.5


def test_增强后的文字图在pdf里是1位黑白(拍照样张, tmp_path) -> None:
    """文字档二值化后存 1-bit：体积小，黑白机打出来也更利。
    get_images() 每项的第 5 个字段是 bpc（每通道位数）。"""
    文字档 = convert.ConvertOptions(enhance=EnhanceOptions(mode=MODE_TEXT, strength=60))
    原样 = convert.ConvertOptions()
    黑白 = convert.images_to_pdf([拍照样张], tmp_path / "增强.pdf", 文字档)
    灰度 = convert.images_to_pdf([拍照样张], tmp_path / "原样.pdf", 原样)

    with pymupdf.open(黑白) as doc:
        assert doc[0].get_images()[0][4] == 1
    with pymupdf.open(灰度) as doc:
        assert doc[0].get_images()[0][4] == 8  # 没增强的走灰度，保留层次
    assert 黑白.stat().st_size < 灰度.stat().st_size


def test_空列表报错而不是生成空pdf(tmp_path) -> None:
    with pytest.raises(ShopPrintError):
        convert.images_to_pdf([], tmp_path / "空.pdf", convert.ConvertOptions())


# ── txt → PDF ───────────────────────────────────────────────────
def test_中文txt能转出可读pdf(tmp_path) -> None:
    src = tmp_path / "通知.txt"
    src.write_text("房屋租赁合同\n甲乙双方经友好协商，达成如下协议。\n", encoding="utf-8")
    out = convert.text_to_pdf(src, tmp_path / "通知.pdf", convert.ConvertOptions())
    with pymupdf.open(out) as doc:
        assert "房屋租赁合同" in doc[0].get_text()


def test_gbk编码的txt也能读(tmp_path) -> None:
    """顾客的 txt 编码很杂，GBK 很常见。乱码打出来就是废纸。"""
    src = tmp_path / "gbk.txt"
    src.write_bytes("承租方：李秀兰".encode("gb18030"))
    out = convert.text_to_pdf(src, tmp_path / "gbk.pdf", convert.ConvertOptions())
    with pymupdf.open(out) as doc:
        assert "李秀兰" in doc[0].get_text()


def test_长文本会自动分页(tmp_path) -> None:
    src = tmp_path / "长.txt"
    src.write_text("\n".join(f"第 {i} 行内容" for i in range(300)), encoding="utf-8")
    out = convert.text_to_pdf(src, tmp_path / "长.pdf", convert.ConvertOptions())
    with pymupdf.open(out) as doc:
        assert doc.page_count > 1


def test_超长的一行会按字宽折行(tmp_path) -> None:
    src = tmp_path / "长行.txt"
    src.write_text("很长的一行" * 60, encoding="utf-8")
    out = convert.text_to_pdf(src, tmp_path / "长行.pdf", convert.ConvertOptions())
    with pymupdf.open(out) as doc:
        page = doc[0]
        for block in page.get_text("blocks"):
            assert block[2] <= page.rect.width + 1  # 没有内容跑到纸外面


# ── PDF 本身 ────────────────────────────────────────────────────
def test_pdf原样通过不重新生成(tmp_path) -> None:
    src = tmp_path / "原始.pdf"
    document = pymupdf.open()
    document.new_page()
    document.save(src)
    document.close()
    assert convert.to_pdf(src) == src


def test_坏的pdf报文件坏了(tmp_path) -> None:
    src = tmp_path / "坏的.pdf"
    src.write_bytes(b"%PDF-1.4 this is not really a pdf")
    with pytest.raises(ShopPrintError) as caught:
        convert.open_pdf(src)
    assert caught.value.kind is ErrorKind.FILE_BROKEN


def test_有密码的pdf提示让顾客重发(tmp_path) -> None:
    src = tmp_path / "加密.pdf"
    document = pymupdf.open()
    document.new_page()
    document.save(src, encryption=pymupdf.PDF_ENCRYPT_AES_256, owner_pw="o", user_pw="u")
    document.close()
    with pytest.raises(ShopPrintError) as caught:
        convert.open_pdf(src)
    assert caught.value.kind is ErrorKind.FILE_ENCRYPTED
    assert "没有密码" in caught.value.friendly


def test_数页数(tmp_path) -> None:
    src = tmp_path / "三页.pdf"
    document = pymupdf.open()
    for _ in range(3):
        document.new_page()
    document.save(src)
    document.close()
    assert convert.page_count(src) == 3


def test_渲染预览出的是png字节(tmp_path) -> None:
    """core 不许 import Qt，所以这里只能返回字节。"""
    src = tmp_path / "一页.pdf"
    document = pymupdf.open()
    document.new_page()
    document.save(src)
    document.close()
    data = convert.render_page_png(src, 0)
    assert data.startswith(b"\x89PNG")


def test_页码越界会被夹回范围(tmp_path) -> None:
    src = tmp_path / "一页.pdf"
    document = pymupdf.open()
    document.new_page()
    document.save(src)
    document.close()
    assert convert.render_page_png(src, 99).startswith(b"\x89PNG")


# ── 缓存 ────────────────────────────────────────────────────────
def test_同一个文件同样选项命中缓存(拍照样张) -> None:
    first = convert.to_pdf(拍照样张)
    stamp = first.stat().st_mtime_ns
    second = convert.to_pdf(拍照样张)
    assert second == first
    assert second.stat().st_mtime_ns == stamp  # 没重新生成


def test_换了选项就是另一份缓存(拍照样张) -> None:
    普通 = convert.to_pdf(拍照样张, convert.ConvertOptions())
    增强 = convert.to_pdf(拍照样张, convert.ConvertOptions(enhance=EnhanceOptions(strength=80)))
    assert 普通 != 增强


def test_内容改了缓存也跟着换(tmp_path) -> None:
    src = 存图(tmp_path / "会变的.png", photographed_document(seed=1))
    第一次 = convert.to_pdf(src)
    存图(src, photographed_document(seed=2))
    assert convert.to_pdf(src) != 第一次


def test_选项签名只认内容不认对象() -> None:
    a = convert.ConvertOptions(enhance=EnhanceOptions(strength=50))
    b = convert.ConvertOptions(enhance=EnhanceOptions(strength=50))
    c = convert.ConvertOptions(enhance=EnhanceOptions(strength=51))
    assert a.signature() == b.signature()
    assert a.signature() != c.signature()


def test_多张图合并的缓存名带张数(拍照样张, tmp_path) -> None:
    second = 存图(tmp_path / "b.png", photographed_document(seed=3))
    path = convert.merged_cache_path([拍照样张, second], convert.ConvertOptions())
    assert "合并2张" in path.name
    第一次 = convert.images_to_pdf_cached([拍照样张, second], convert.ConvertOptions())
    assert convert.images_to_pdf_cached([拍照样张, second], convert.ConvertOptions()) == 第一次


# ── Office（要装 Office）─────────────────────────────────────────
@pytest.mark.needs_office
def test_word转pdf(tmp_path) -> None:
    from docx import Document

    src = tmp_path / "合同.docx"
    document = Document()
    document.add_paragraph("房屋租赁合同")
    document.save(src)

    out = convert.to_pdf(src)
    with pymupdf.open(out) as doc:
        assert doc.page_count >= 1


@pytest.mark.needs_office
def test_excel宽表不会被切成几十页(tmp_path) -> None:
    """复印店最常见的浪费纸事故。必须 FitToPagesWide=1。"""
    from openpyxl import Workbook

    src = tmp_path / "宽表.xlsx"
    book = Workbook()
    sheet = book.active
    for column in range(1, 40):
        sheet.cell(row=1, column=column, value=f"第{column}列")
        sheet.cell(row=2, column=column, value=column * 100)
    book.save(src)

    out = convert.to_pdf(src)
    with pymupdf.open(out) as doc:
        assert doc.page_count == 1
